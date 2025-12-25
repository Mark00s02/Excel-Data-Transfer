import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import re

# ===== KEYWORDS =====
KEYWORDS = ["no cng", "no zenput", "no dsd", "no eod", "emailed", "missing cng", 
            "missing zenput", "missing dsd", "missing eod"]

# ===== COLORS =====
PRIMARY = "#4f46e5"      # Indigo
SECONDARY = "#22c55e"    # Green
ACCENT = "#ec4899"       # Pink
BG = "#f8fafc"           # Light gray
CARD = "#ffffff"

# ===== COLUMNS =====
CNG_COL = "CNG Docs"
ZENPUT_COL = "Zenput Docs"

class ExcelTransferApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Transfer System")
        self.root.geometry("820x600")
        self.root.configure(bg=BG)

        self.source_file = ""
        self.dest_file = ""

        self.source_columns = []
        self.dest_columns = []

        self.mappings = {}

        self.build_ui()

    # ===== TIME DETECTION =====
    def is_time_value(self, value):
        if pd.isna(value):
            return False
        return bool(re.search(r"\d{1,2}:\d{2}", str(value)))

    # ===== KEYWORD DETECTION =====
    def contains_any_keyword(self, value):
        """Return True if the cell contains any keyword."""
        if pd.isna(value):
            return False
        text = str(value).lower()
        tokens = re.split(r"[;,]", text)
        tokens = [t.strip() for t in tokens if t.strip()]
        tokens = [t for t in tokens if not self.is_time_value(t)]
        return any(any(k in t for k in KEYWORDS) for t in tokens)

    # ===== EXTRACT TEXT ONLY =====
    def extract_text_only(self, value):
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if self.is_time_value(text):
            return ""
        return text

    # ===== UI BUILD =====
    def build_ui(self):
        # HEADER
        header = tk.Frame(self.root, bg=PRIMARY, height=70)
        header.pack(fill="x")
        tk.Label(
            header,
            text="📊 Excel Data Transfer System",
            bg=PRIMARY,
            fg="white",
            font=("Segoe UI", 20, "bold")
        ).pack(pady=15)

        # MAIN CONTAINER
        container = tk.Frame(self.root, bg=BG)
        container.pack(fill="both", expand=True, padx=20, pady=15)

        # FILE SELECTION
        file_card = tk.Frame(container, bg=CARD, bd=0, relief="flat")
        file_card.pack(fill="x", pady=10)
        tk.Label(file_card, text="📂 Select Excel Files", bg=CARD, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=15, pady=10)

        btn_frame = tk.Frame(file_card, bg=CARD)
        btn_frame.pack(padx=15, pady=10)
        tk.Button(btn_frame, text="Select Source Excel", bg=PRIMARY, fg="white", font=("Segoe UI", 10, "bold"),
                  padx=15, pady=8, relief="flat", command=self.load_source).grid(row=0, column=0, padx=10)
        tk.Button(btn_frame, text="Select Destination Excel", bg=ACCENT, fg="white", font=("Segoe UI", 10, "bold"),
                  padx=15, pady=8, relief="flat", command=self.load_dest).grid(row=0, column=1, padx=10)

        # ===== SCROLLABLE MAPPING CARD =====
        mapping_card = tk.LabelFrame(
            container,
            text="🔗 Column Mapping",
            bg=CARD,
            font=("Segoe UI", 12, "bold"),
            labelanchor="n"
        )
        mapping_card.pack(fill="both", expand=True, pady=15)

        # Canvas + Scrollbar
        canvas = tk.Canvas(mapping_card, bg=CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(mapping_card, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Inner frame (THIS is where widgets go)
        self.mapping_frame = tk.Frame(canvas, bg=CARD)
        canvas.create_window((0, 0), window=self.mapping_frame, anchor="nw")

        # Auto-resize scroll region
        def on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        self.mapping_frame.bind("<Configure>", on_configure)

        # Mouse wheel support (Windows)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))


        # PROGRESS BAR
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(self.root, variable=self.progress_var, maximum=100)
        self.progress.pack(fill="x", padx=40, pady=(5, 0))
        self.status_label = tk.Label(self.root, text="Ready", bg=BG, fg="gray", font=("Segoe UI", 9))
        self.status_label.pack(pady=(2, 10))

        # ACTION BUTTON
        tk.Button(self.root, text="🚀 Run Transfer", bg=SECONDARY, fg="white", font=("Segoe UI", 12, "bold"),
                  padx=25, pady=10, relief="flat", command=self.run_transfer).pack(pady=15)

        # FOOTER
        footer = tk.Label(self.root, text="Detects: no cng | no zenput | no dsd | no eod | emailed", bg=BG, fg="gray", font=("Segoe UI", 9))
        footer.pack(pady=5)

    # LOAD SOURCE
    def load_source(self):
        self.source_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not self.source_file:
            return
        df = pd.read_excel(self.source_file)
        self.source_columns = list(df.columns)
        self.refresh_mapping_ui()

    # LOAD DESTINATION
    def load_dest(self):
        self.dest_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not self.dest_file:
            return
        wb = load_workbook(self.dest_file)
        ws = wb.active
        self.dest_columns = [cell.value for cell in ws[1]]
        self.refresh_mapping_ui()

    # REFRESH MAPPING UI
    def refresh_mapping_ui(self):
        for widget in self.mapping_frame.winfo_children():
            widget.destroy()
        if not self.source_columns or not self.dest_columns:
            return

        tk.Label(self.mapping_frame, text="Source Column", bg=CARD, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=15, pady=8)
        tk.Label(self.mapping_frame, text="Destination Column", bg=CARD, font=("Segoe UI", 10, "bold")).grid(row=0, column=1, padx=15, pady=8)

        self.mappings.clear()
        for i, src_col in enumerate(self.source_columns):
            tk.Label(self.mapping_frame, text=src_col, bg=CARD, anchor="w", font=("Segoe UI", 10)).grid(row=i + 1, column=0, sticky="w", padx=15, pady=4)
            combo = ttk.Combobox(self.mapping_frame, values=["-- Don't Add --"] + self.dest_columns, state="readonly", width=30)
            combo.set("-- Don't Add --")
            combo.grid(row=i + 1, column=1, padx=15, pady=4)
            self.mappings[src_col] = combo

    # RUN TRANSFER
    def run_transfer(self):
        if not self.source_file or not self.dest_file:
            messagebox.showerror("Error", "Please select both Excel files.")
            return

        source_df = pd.read_excel(self.source_file)
        wb = load_workbook(self.dest_file)
        ws = wb.active
        dest_headers = [cell.value for cell in ws[1]]

        # 🔑 Determine what the user actually mapped
        zenput_mapped = (
            ZENPUT_COL in self.mappings and
            self.mappings[ZENPUT_COL].get() != "-- Don't Add --"
        )

        cng_mapped = (
            CNG_COL in self.mappings and
            self.mappings[CNG_COL].get() != "-- Don't Add --"
        )

        total_rows = len(source_df)
        self.progress_var.set(0)
        self.status_label.config(text="Processing...")

        rows_added = 0

        for idx, row in source_df.iterrows():
            matched_keywords = []

            # ✅ Check Zenput ONLY if user mapped it
            if zenput_mapped and ZENPUT_COL in source_df.columns:
                if self.contains_any_keyword(row[ZENPUT_COL]):
                    text = self.extract_text_only(row[ZENPUT_COL])
                    if text:
                        matched_keywords.append(text)

            # ✅ Check CNG ONLY if user mapped it
            if cng_mapped and CNG_COL in source_df.columns:
                if self.contains_any_keyword(row[CNG_COL]):
                    text = self.extract_text_only(row[CNG_COL])
                    if text:
                        matched_keywords.append(text)

            # ❌ Skip if nothing matched
            if not matched_keywords:
                continue

            # ✅ Build destination row
            new_row = {}
            for src_col, combo in self.mappings.items():
                dest_col = combo.get()
                if dest_col == "-- Don't Add --":
                    continue

                if src_col in [CNG_COL, ZENPUT_COL]:
                    value = self.extract_text_only(row[src_col])
                    if value:
                        if dest_col in new_row:
                            new_row[dest_col] += ", " + value
                        else:
                            new_row[dest_col] = value
                else:
                    new_row[dest_col] = row[src_col]

            # ✅ Write to Excel
            write_row = ws.max_row + 1
            for col_idx, header in enumerate(dest_headers, start=1):
                if header in new_row:
                    ws.cell(row=write_row, column=col_idx, value=new_row[header])

            rows_added += 1

            # Progress update
            percent = ((idx + 1) / total_rows) * 100
            self.progress_var.set(percent)
            self.status_label.config(text=f"Processing row {idx + 1} of {total_rows}")
            self.root.update_idletasks()

        wb.save(self.dest_file)
        self.progress_var.set(100)
        self.status_label.config(text="Completed")
        messagebox.showinfo("Success", f"🎉 {rows_added} rows inserted successfully!")

# ==============================
# RUN APP
# ==============================
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelTransferApp(root)
    root.mainloop()
